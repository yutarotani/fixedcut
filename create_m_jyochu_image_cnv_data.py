#!/usr/bin/env python3
"""
m_jyochu_image_cnv テーブルへ Excel データを投入するスクリプト。
"""

import os
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(__file__))

from fixedcut_app import app, db
from fixedcut_app.models.m_jyochu_image_cnv import MJyochuImageCnv


def to_text(value):
    if value is None:
        return ""
    return str(value).strip()


def to_int(value, default=0):
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def import_excel_data() -> bool:
    excel_path = Path(__file__).parent / "fixedcut_app" / "templates" / "static" / "xlsx" / "m_jyochu_image_cnv" / "現データ.xlsx"

    if not excel_path.exists():
        print(f"❌ Excelファイルが見つかりません: {excel_path}")
        return False

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    with app.app_context():
        print(f"開始: {excel_path}")
        print(f"シート: {ws.title}, 行数: {ws.max_row}, 列数: {ws.max_column}")

        inserted = 0
        updated = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=16, values_only=True):
            fixed_cut_id = to_text(row[1])
            if not fixed_cut_id:
                skipped += 1
                continue

            record = db.session.get(MJyochuImageCnv, fixed_cut_id)
            if record is None:
                record = MJyochuImageCnv(fixed_cut_id=fixed_cut_id)
                db.session.add(record)
                inserted += 1
            else:
                updated += 1

            record.data_description = to_text(row[0])
            record.fixed_cut_img_explanation = to_text(row[2])
            record.upd_count = to_int(row[3], 0)
            record.created_datetime = to_text(row[4]) or "CURRENT_TIMESTAMP"
            record.created_user = to_text(row[5]) or "Initial"
            record.created_term = to_text(row[6])
            record.created_pgm = to_text(row[7])
            record.created_trn_id = to_text(row[8])
            record.updated_datetime = to_text(row[9]) or "CURRENT_TIMESTAMP"
            record.updated_user = to_text(row[10]) or "Initial"
            record.updated_term = to_text(row[11])
            record.updated_pgm = to_text(row[12])
            record.updated_trn_id = to_text(row[13])
            record.patch_no = to_text(row[14])
            record.patch_datetime = to_text(row[15]) or "[NULL]"

        try:
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            print(f"❌ DBコミットに失敗しました: {exc}")
            return False

        total = db.session.query(MJyochuImageCnv).count()
        print("✅ 取り込み完了")
        print(f"Inserted: {inserted}")
        print(f"Updated: {updated}")
        print(f"Skipped: {skipped}")
        print(f"Total rows in table: {total}")

    return True


if __name__ == "__main__":
    ok = import_excel_data()
    if not ok:
        sys.exit(1)
