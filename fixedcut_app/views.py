from flask import render_template, request, redirect, url_for, flash, send_from_directory, send_file
from datetime import datetime
from fixedcut_app import app, db
from fixedcut_app.models.fixedcut import FixedCut
from fixedcut_app.models.m_jyochu_image_cnv import MJyochuImageCnv
from fixedcut_app.models.senkyo_person import SenkyoPerson
from fixedcut_app.models.senkyo_sendgroup import SenkyoSendGroup
import os, pathlib, sqlite3, shutil
import importlib
from sqlalchemy import and_
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from io import BytesIO
import csv
import unicodedata



@app.route('/')
def index():
    from fixedcut_app.views_index import index_handler
    return index_handler()


def _to_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_fixedcut_id(value):
    # 全角英数を半角へ正規化し、前後空白を除去する。
    return unicodedata.normalize('NFKC', _to_text(value))


def _normalize_area_text(value):
    # area照合のぶれを減らすため、全角/半角スペースを除去して比較する。
    return _to_text(value).replace('\u3000', '').replace(' ', '')


def _to_int(value, default=0):
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _xlsx_base_dir():
    return pathlib.Path(app.static_folder) / 'xlsx'


def _senkyo_data_extensions():
    return {'.xlsx', '.xlsm', '.xls', '.csv'}


def _normalize_senkyoku(value):
    text = _to_text(value)
    for ch in ('\u3000', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '０', '１', '２', '３', '４', '５', '６', '７', '８', '９'):
        text = text.replace(ch, '')
    return text.strip().strip('区')


def _to_int_or_none(value):
    text = _to_text(value)
    if text == '':
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _parse_date_or_none(value):
    text = _to_text(value).strip()
    if text == '':
        return None
    for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y.%m.%d', '%Y%m%d'):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _parse_datetime_local_or_none(value):
    text = _to_text(value)
    if text == '':
        return None
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M'):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _match_sendgroup_value(syubetu, senkyoku, senkyoku_no_num, hirei):
    rows = db.session.query(SenkyoSendGroup).order_by(SenkyoSendGroup.id.asc()).all()
    if not rows:
        return ''

    area_candidates = []
    for raw in (senkyoku, hirei):
        normalized = _normalize_area_text(raw)
        if normalized and normalized not in area_candidates:
            area_candidates.append(normalized)

    # 1) area + 小選挙区番号 の一致を優先
    if senkyoku_no_num is not None:
        for area in area_candidates:
            for row in rows:
                if _normalize_area_text(row.area) == area and row.syosenkyoNum == senkyoku_no_num:
                    return _to_text(row.sendGroup)

    # 2) area の一致で決定（要件: SenkyoSendGroup.area を参照して sendGroup を設定）
    for area in area_candidates:
        for row in rows:
            if _normalize_area_text(row.area) == area:
                return _to_text(row.sendGroup)

    return ''


def _iter_upload_rows(file_path):
    ext = file_path.suffix.lower()
    if ext == '.csv':
        rows = None
        for enc in ('cp932', 'shift_jis', 'utf-8-sig', 'utf-8'):
            try:
                with open(file_path, 'r', encoding=enc, newline='') as f:
                    rows = list(csv.reader(f))
                break
            except UnicodeDecodeError:
                continue
        if rows is None:
            raise ValueError('CSVの文字コードを判定できませんでした')
        for row in rows[1:]:
            yield row
        return

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        yield list(row)


def _upsert_senkyo_person_from_cd_file(file_path):
    inserted = 0
    updated = 0
    skipped = 0

    for row in _iter_upload_rows(file_path):
        if row is None or len(row) < 16:
            skipped += 1
            continue

        personid = _to_int_or_none(row[5])
        if personid is None:
            skipped += 1
            continue

        syubetu = _to_text(row[1])
        # 種別文字列に依存せず、候補エリアとして選挙区列を利用する。
        senkyoku = _normalize_senkyoku(row[3])

        senkyoku_no_num = _to_int_or_none(row[2])
        if senkyoku_no_num is None:
            senkyoku_no = ''
        else:
            senkyoku_no = f'{senkyoku_no_num}区'

        hirei = _to_text(row[4]).replace('\u3000', '')
        send_group = _match_sendgroup_value(syubetu, senkyoku, senkyoku_no_num, hirei)

        payload = {
            'syubetu': syubetu,
            'senkyoku': senkyoku,
            'senkyokuNo': senkyoku_no,
            'sendGroup': send_group,
            'hirei': hirei,
            'name': _to_text(row[6]).replace('\u3000', ''),
            'hurigana': _to_text(row[7]).replace('\u3000', ''),
            'name_jikai': _to_text(row[8]).replace('\u3000', ''),
            'kyodo_name': _to_text(row[9]).replace('\u3000', ''),
            'seibetsu': _to_text(row[10]),
            'seito': _to_text(row[11]).replace('\u3000', ''),
            'genshinbetu': _to_text(row[12]),
            'facefilename': _to_text(row[13]),
            'photo_date': _parse_date_or_none(row[15]),
            'CD_No': _to_text(row[14]),
            'fixedcutID': '',
            'updateCount': 0,
        }

        record = db.session.query(SenkyoPerson).filter(SenkyoPerson.id == personid).first()
        if record is None:
            record = SenkyoPerson(id=personid, **payload)
            db.session.add(record)
            inserted += 1
        else:
            for key, value in payload.items():
                setattr(record, key, value)
            updated += 1

    return inserted, updated, skipped


def _read_tabular_file(file_path):
    ext = file_path.suffix.lower()

    if ext == '.csv':
        rows = None
        for enc in ('cp932', 'shift_jis', 'utf-8-sig', 'utf-8'):
            try:
                with open(file_path, 'r', encoding=enc, newline='') as f:
                    rows = list(csv.reader(f))
                break
            except UnicodeDecodeError:
                continue
        if rows is None:
            raise ValueError('CSVの文字コードを判定できませんでした')
        if not rows:
            return [], []
        return rows[0], rows[1:]

    if ext == '.xls':
        try:
            xlrd = importlib.import_module('xlrd')
        except ImportError as exc:
            raise ValueError('.xls形式の読込に必要なライブラリ(xlrd)が未導入です') from exc

        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        if ws.nrows == 0:
            return [], []
        headers = [ws.cell_value(0, c) for c in range(ws.ncols)]
        data_rows = [ws.row_values(r) for r in range(1, ws.nrows)]
        return headers, data_rows

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    if ws.max_row == 0:
        return [], []
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    data_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True))
    return headers, data_rows


def _import_senkyo_sendgroup_from_file(file_path):
    headers, data_rows = _read_tabular_file(file_path)
    if not headers:
        return 0, 0

    normalized_headers = [_to_text(h) for h in headers]
    header_index = {h: i for i, h in enumerate(normalized_headers)}

    def pick(row, candidates, fallback_index):
        for name in candidates:
            idx = header_index.get(name)
            if idx is not None and idx < len(row):
                return row[idx]
        if fallback_index < len(row):
            return row[fallback_index]
        return None

    inserted = 0
    skipped = 0

    # 取込は毎回全置換にして、ファイル内容とDBを一致させる。
    db.session.query(SenkyoSendGroup).delete(synchronize_session=False)

    for row in data_rows:
        if row is None:
            skipped += 1
            continue

        syubetu = _to_text(pick(row, ('選挙種別', '種別', 'syubetu'), 0))
        area = _to_text(pick(row, ('エリア名', '選挙区', 'area'), 1))
        syosenkyo_num = _to_int_or_none(pick(row, ('小選挙区数', '小選挙数', '小選挙区', 'syosenkyoNum'), 2))
        if syosenkyo_num is None:
            syosenkyo_num = 0
        send_group = _to_text(pick(row, ('共同送信グループ', '配信グループ', 'sendGroup'), 3))

        if not syubetu and not area and syosenkyo_num is None and not send_group:
            skipped += 1
            continue

        record = SenkyoSendGroup(
            syubetu=syubetu,
            area=area,
            syosenkyoNum=syosenkyo_num,
            sendGroup=send_group,
        )
        db.session.add(record)
        inserted += 1

    return inserted, skipped


@app.route('/m_jyochu_image_cnv/upload', methods=['POST'])
def upload_m_jyochu_excel():
    upload_file = request.files.get('excelFile')
    if not upload_file or upload_file.filename == '':
        flash('※Excelファイルを選択してください※')
        return redirect(url_for('index'))

    ext = pathlib.Path(upload_file.filename).suffix.lower()
    if ext not in {'.xlsx', '.xlsm', '.xls', '.csv'}:
        flash('※Excel/CSV形式(.xlsx/.xlsm/.xls/.csv)のファイルを選択してください※')
        return redirect(url_for('index'))

    xlsx_dir = _xlsx_base_dir()
    xlsx_dir.mkdir(parents=True, exist_ok=True)

    safe_name = secure_filename(upload_file.filename)
    # 日本語ファイル名などで secure_filename が拡張子を落とすことがあるため、拡張子は必ず維持する。
    if (not safe_name) or (pathlib.Path(safe_name).suffix.lower() != ext):
        safe_name = f'uploaded_m_jyochu_image_cnv{ext}'

    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    save_path = xlsx_dir / f'{timestamp}_{safe_name}'
    upload_file.save(save_path)

    if ext == '.csv':
        rows = None
        for enc in ('utf-8-sig', 'cp932', 'shift_jis'):
            try:
                with open(save_path, 'r', encoding=enc, newline='') as f:
                    rows = list(csv.reader(f))
                break
            except UnicodeDecodeError:
                continue
        if rows is None:
            flash('※CSVの文字コード判定に失敗しました※')
            return redirect(url_for('index'))
        if len(rows) < 2:
            flash('※データ行がありません※')
            return redirect(url_for('index'))

        headers = [_to_text(x) for x in rows[0]]
        data_rows = rows[1:]
    elif ext == '.xls':
        try:
            xlrd = importlib.import_module('xlrd')
        except ImportError:
            flash('※.xls形式の読込に必要なライブラリ(xlrd)が未導入です。Excelで .xlsx か .csv に変換して再度取り込んでください※')
            return redirect(url_for('index'))

        try:
            wb = xlrd.open_workbook(save_path)
            ws = wb.sheet_by_index(0)
        except Exception as exc:
            flash(f'※Excel読込に失敗しました※ {exc}')
            return redirect(url_for('index'))

        if ws.nrows < 2:
            flash('※データ行がありません※')
            return redirect(url_for('index'))

        headers = [_to_text(ws.cell_value(0, c)) for c in range(ws.ncols)]
        data_rows = [ws.row_values(r) for r in range(1, ws.nrows)]
    else:
        try:
            wb = load_workbook(save_path, data_only=True)
            ws = wb.active
        except Exception as exc:
            flash(f'※Excel読込に失敗しました※ {exc}')
            return redirect(url_for('index'))

        if ws.max_row < 2:
            flash('※データ行がありません※')
            return redirect(url_for('index'))

        headers = [_to_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        data_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max(ws.max_column, 16), values_only=True))

    header_index = {h: i for i, h in enumerate(headers)}

    def pick(row, header_name, fallback_index):
        idx = header_index.get(header_name)
        if idx is not None and idx < len(row):
            return row[idx]
        if fallback_index < len(row):
            return row[fallback_index]
        return None

    # 取込はアップロードファイルの行順をそのまま処理する（ソートしない）。
    inserted = 0
    updated = 0
    unchanged = 0
    skipped = 0

    try:
        for row in data_rows:
            fixed_cut_id = _to_text(pick(row, 'fixed_cut_id', 1))
            if not fixed_cut_id:
                skipped += 1
                continue

            new_values = {
                'data_description': _to_text(pick(row, 'データ説明', 0)),
                'fixed_cut_img_explanation': _to_text(pick(row, 'fixed_cut_img_explanation', 2)),
                'upd_count': _to_int(pick(row, 'upd_count', 3), 0),
                'created_datetime': _to_text(pick(row, 'created_datetime', 4)) or 'CURRENT_TIMESTAMP',
                'created_user': _to_text(pick(row, 'created_user', 5)) or 'Initial',
                'created_term': _to_text(pick(row, 'created_term', 6)),
                'created_pgm': _to_text(pick(row, 'created_pgm', 7)),
                'created_trn_id': _to_text(pick(row, 'created_trn_id', 8)),
                'updated_datetime': _to_text(pick(row, 'updated_datetime', 9)) or 'CURRENT_TIMESTAMP',
                'updated_user': _to_text(pick(row, 'updated_user', 10)) or 'Initial',
                'updated_term': _to_text(pick(row, 'updated_term', 11)),
                'updated_pgm': _to_text(pick(row, 'updated_pgm', 12)),
                'updated_trn_id': _to_text(pick(row, 'updated_trn_id', 13)),
                'patch_no': _to_text(pick(row, 'patch_no', 14)),
                'patch_datetime': _to_text(pick(row, 'patch_datetime', 15)) or '[NULL]',
            }

            record = db.session.get(MJyochuImageCnv, fixed_cut_id)
            if record is None:
                record = MJyochuImageCnv(fixed_cut_id=fixed_cut_id, **new_values)
                db.session.add(record)
                inserted += 1
                continue

            changed = False
            for key, val in new_values.items():
                if getattr(record, key) != val:
                    setattr(record, key, val)
                    changed = True

            if changed:
                updated += 1
            else:
                unchanged += 1

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        flash(f'※差分反映に失敗しました※ {exc}')
        return redirect(url_for('index'))

    flash('Excel取込と差分反映が完了しました')
    flash(f'保存先: {save_path}')
    flash(f'追加: {inserted}件 / 更新: {updated}件 / 変更なし: {unchanged}件 / スキップ: {skipped}件')
    return redirect(url_for('index'))


@app.route('/m_jyochu_image_cnv/download', methods=['GET'])
def download_m_jyochu_excel():
    # ダウンロードはDBの取得順をそのまま使う（ソートしない）。
    records = db.session.query(MJyochuImageCnv).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'm_jyochu_image_cnv'

    ws.append([
        'データ説明',
        'fixed_cut_id',
        'fixed_cut_img_explanation',
        'upd_count',
        'created_datetime',
        'created_user',
        'created_term',
        'created_pgm',
        'created_trn_id',
        'updated_datetime',
        'updated_user',
        'updated_term',
        'updated_pgm',
        'updated_trn_id',
        'patch_no',
        'patch_datetime',
    ])

    for rec in records:
        ws.append([
            rec.data_description,
            rec.fixed_cut_id,
            rec.fixed_cut_img_explanation,
            rec.upd_count,
            rec.created_datetime,
            rec.created_user,
            rec.created_term,
            rec.created_pgm,
            rec.created_trn_id,
            rec.updated_datetime,
            rec.updated_user,
            rec.updated_term,
            rec.updated_pgm,
            rec.updated_trn_id,
            rec.patch_no,
            rec.patch_datetime,
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"m_jyochu_image_cnv_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/m_jyochu_image_cnv/reset', methods=['POST'])
def reset_m_jyochu_image_cnv_data():
    m_jyochu_dir = _xlsx_base_dir() / 'm_jyochu_image_cnv'
    deleted_files = 0

    try:
        if m_jyochu_dir.exists():
            for file_path in m_jyochu_dir.glob('*'):
                if file_path.is_file():
                    file_path.unlink()
                    deleted_files += 1

        deleted_rows = db.session.query(MJyochuImageCnv).delete(synchronize_session=False)
        db.session.commit()
        flash('m_jyochu_image_cnvのフォルダ内ファイル削除と全削除を実行しました')
        flash(f'削除ファイル数: {deleted_files}件 / 削除レコード数: {deleted_rows}件')
    except Exception as exc:
        db.session.rollback()
        flash(f'※m_jyochu_image_cnv削除処理に失敗しました※ {exc}')

    return redirect(url_for('index'))


@app.route('/senkyo')
def senkyo():
    from fixedcut_app.views_senkyo import senkyo_handler
    return senkyo_handler()


@app.route('/senkyo/sendgroup_text')
def senkyo_sendgroup_text():
    from fixedcut_app.views_senkyo import senkyo_sendgroup_text_handler
    return senkyo_sendgroup_text_handler()


def _save_excel_to_static_subdir(upload_file, sub_dir_name, default_name):
    if not upload_file or upload_file.filename == '':
        return False, '※Excelファイルを選択してください※'

    ext = pathlib.Path(upload_file.filename).suffix.lower()
    if ext not in _senkyo_data_extensions():
        return False, '※Excel/CSV形式(.xlsx/.xlsm/.xls/.csv)のファイルを選択してください※'

    base_dir = _xlsx_base_dir() / sub_dir_name
    base_dir.mkdir(parents=True, exist_ok=True)

    safe_name = secure_filename(upload_file.filename)
    # 日本語名などで secure_filename が拡張子を落とすことがあるため、拡張子は必ず維持する。
    if (not safe_name) or (pathlib.Path(safe_name).suffix.lower() != ext):
        safe_name = f'uploaded_{sub_dir_name}{ext}'
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    save_path = base_dir / f'{timestamp}_{safe_name}'
    upload_file.save(save_path)

    return True, f'保存完了: {save_path}'


@app.route('/senkyo/upload_cd', methods=['POST'])
def upload_senkyo_cd_excel():
    upload_file = request.files.get('cdExcelFile')
    ok, message = _save_excel_to_static_subdir(upload_file, 'CD', 'cd_upload.xlsx')
    flash(message)

    if ok:
        saved_path = pathlib.Path(message.split(': ', 1)[1])
        try:
            inserted, updated, skipped = _upsert_senkyo_person_from_cd_file(saved_path)
            db.session.commit()
            flash(f'SenkyoPerson 追加: {inserted}件 / 更新: {updated}件 / スキップ: {skipped}件')
        except Exception as exc:
            db.session.rollback()
            flash(f'※SenkyoPerson取込に失敗しました※ {exc}')

    return redirect(url_for('senkyo'))


@app.route('/senkyo/upload_sendgroup', methods=['POST'])
def upload_senkyo_sendgroup_excel():
    upload_file = request.files.get('sendgroupExcelFile')
    ok, message = _save_excel_to_static_subdir(upload_file, 'sendgroup', 'sendgroup_upload.xlsx')
    flash(message)

    if ok:
        saved_path = pathlib.Path(message.split(': ', 1)[1])
        try:
            inserted, skipped = _import_senkyo_sendgroup_from_file(saved_path)
            db.session.commit()
            flash(f'SenkyoSendGroup 取込完了: {inserted}件 / スキップ: {skipped}件')
        except Exception as exc:
            db.session.rollback()
            flash(f'※SenkyoSendGroup取込に失敗しました※ {exc}')

    return redirect(url_for('senkyo'))


def _build_excel_response(records, sheet_title, file_prefix):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    if records:
        columns = [c.name for c in records[0].__table__.columns]
    else:
        columns = []

    if columns:
        ws.append(columns)
        for rec in records:
            ws.append([getattr(rec, col) for col in columns])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"{file_prefix}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/senkyo/download_person_excel', methods=['GET'])
def download_senkyo_person_excel():
    records = db.session.query(SenkyoPerson).order_by(SenkyoPerson.id.asc()).all()
    return _build_excel_response(records, 'SenkyoPerson', 'senkyo_person')


@app.route('/senkyo/download_sendgroup_excel', methods=['GET'])
def download_senkyo_sendgroup_excel():
    records = db.session.query(SenkyoSendGroup).order_by(SenkyoSendGroup.id.asc()).all()
    return _build_excel_response(records, 'SenkyoSendGroup', 'senkyo_sendgroup')


@app.route('/senkyo/reset_cd_person', methods=['POST'])
def reset_cd_person_data():
    cd_dir = _xlsx_base_dir() / 'CD'
    deleted_files = 0

    try:
        if cd_dir.exists():
            for file_path in cd_dir.glob('*'):
                if file_path.is_file() and file_path.suffix.lower() in _senkyo_data_extensions():
                    file_path.unlink()
                    deleted_files += 1

        deleted_rows = db.session.query(SenkyoPerson).delete(synchronize_session=False)
        db.session.commit()

        flash('CDフォルダのExcel削除とSenkyoPerson全削除を実行しました')
        flash(f'削除ファイル数: {deleted_files}件 / 削除レコード数: {deleted_rows}件')
    except Exception as exc:
        db.session.rollback()
        flash(f'※削除処理に失敗しました※ {exc}')

    return redirect(url_for('senkyo'))


@app.route('/senkyo/reset_sendgroup', methods=['POST'])
def reset_sendgroup_data():
    sendgroup_dir = _xlsx_base_dir() / 'sendgroup'
    deleted_files = 0

    try:
        if sendgroup_dir.exists():
            for file_path in sendgroup_dir.glob('*'):
                if file_path.is_file():
                    file_path.unlink()
                    deleted_files += 1

        deleted_rows = db.session.query(SenkyoSendGroup).delete(synchronize_session=False)
        db.session.commit()
        flash('sendgroupフォルダ内ファイル削除とSenkyoSendGroup全削除を実行しました')
        flash(f'削除ファイル数: {deleted_files}件 / 削除レコード数: {deleted_rows}件')
    except Exception as exc:
        db.session.rollback()
        flash(f'※SenkyoSendGroup削除処理に失敗しました※ {exc}')

    return redirect(url_for('senkyo'))


@app.route('/senkyo_person_table', methods=['GET', 'POST'])
@app.route('/senkyo_person_table/<int:page>', methods=['GET', 'POST'])
def senkyo_person_table(page=1):
    per_page = 100

    if request.method == 'POST' and request.form.get('action') == 'bulk_update':
        updated = 0
        invalid = 0
        fixedcut_candidates = {}
        allow_fixedcut_insert = request.form.get('allow_fixedcut_insert') == '1'

        person_ids = request.form.getlist('person_ids')
        for person_id_text in person_ids:
            person_id = _to_int_or_none(person_id_text)
            if person_id is None:
                continue

            rec = db.session.get(SenkyoPerson, person_id)
            if rec is None:
                continue

            changed = False

            fixedcut_id_input = _normalize_fixedcut_id(request.form.get(f'fixedcutID_{person_id}', ''))
            if fixedcut_id_input != '' and rec.fixedcutID != fixedcut_id_input:
                rec.fixedcutID = fixedcut_id_input
                changed = True

            men_name_input = _to_text(request.form.get(f'MenName_{person_id}', ''))
            if men_name_input != '' and rec.MenName != men_name_input:
                rec.MenName = men_name_input
                changed = True

            operater_input = _to_text(request.form.get(f'operater_{person_id}', ''))
            if operater_input != '' and rec.operater != operater_input:
                rec.operater = operater_input
                changed = True

            output_flg_input = request.form.get(f'output_Flg_{person_id}') == '1'
            current_output_flg = bool(rec.output_Flg)
            if current_output_flg != output_flg_input:
                rec.output_Flg = output_flg_input
                changed = True

            store_date_text = _to_text(request.form.get(f'store_date_{person_id}', ''))
            if store_date_text != '':
                parsed_store_date = _parse_datetime_local_or_none(store_date_text)
                if parsed_store_date is None:
                    invalid += 1
                elif rec.store_date != parsed_store_date:
                    rec.store_date = parsed_store_date
                    changed = True

            if changed:
                updated += 1

            fixedcut_id_for_sync = _normalize_fixedcut_id(request.form.get(f'fixedcutID_{person_id}', ''))
            if fixedcut_id_for_sync == '':
                fixedcut_id_for_sync = _to_text(rec.fixedcutID)

            if fixedcut_id_for_sync:
                fixedcut_candidates[fixedcut_id_for_sync] = {
                    'midashi': _to_text(rec.name),
                    'prodFlg': bool(_to_text(rec.MenName) and _to_text(rec.operater) and rec.store_date is not None),
                }

        existing_fixedcut_ids = []
        inserted_fixedcut = 0
        if fixedcut_candidates:
            for fixedcut_id, payload in fixedcut_candidates.items():
                exists = db.session.query(FixedCut.id).filter(FixedCut.id == fixedcut_id).first() is not None
                if exists:
                    existing_fixedcut_ids.append(fixedcut_id)
                    continue

                if allow_fixedcut_insert:
                    db.session.add(FixedCut(
                        id=fixedcut_id,
                        midashi=payload['midashi'],
                        Str='',
                        colorUrl='',
                        monoUrl='',
                        GWFlg=False,
                        prodFlg=payload['prodFlg'],
                        OTFlg=False,
                        comment='選挙顔写真固定カット',
                    ))
                    inserted_fixedcut += 1

        try:
            db.session.commit()
            flash(f'SenkyoPerson 差分更新完了: {updated}件')
            if invalid > 0:
                flash(f'日時形式が不正で更新しなかった件数: {invalid}件')
            if existing_fixedcut_ids:
                flash(f'FixedCut既存ID: {len(existing_fixedcut_ids)}件')
            if inserted_fixedcut > 0:
                flash(f'FixedCut新規追加: {inserted_fixedcut}件')
        except Exception as exc:
            db.session.rollback()
            flash(f'※SenkyoPerson差分更新に失敗しました※ {exc}')
        # 更新後は同一リクエスト内で再検索して最新状態を表示する。

    res = {
        "personID": request.values.get("personID", ""),
        "syubetsu": request.values.get("syubetsu", ""),
        "senkyoku": request.values.get("senkyoku", ""),
        "senkyokuNo": request.values.get("senkyokuNo", ""),
        "sendgroup": request.values.get("sendgroup", ""),
        "hirei": request.values.get("hirei", ""),
        "shimei": request.values.get("shimei", ""),
        "CD_No": request.values.get("CD_No", ""),
        "startdate": request.values.get("startdate", ""),
        "enddate": request.values.get("enddate", ""),
    }

    savelist = [
        res["personID"],
        res["syubetsu"],
        res["senkyoku"],
        res["senkyokuNo"],
        res["sendgroup"],
        res["hirei"],
        res["shimei"],
        res["CD_No"],
        res["startdate"],
        res["enddate"],
    ]

    query = db.session.query(
        SenkyoPerson.id,
        SenkyoPerson.syubetu,
        SenkyoPerson.senkyoku,
        SenkyoPerson.senkyokuNo,
        SenkyoPerson.sendGroup,
        SenkyoPerson.hirei,
        SenkyoPerson.name,
        SenkyoPerson.hurigana,
        SenkyoPerson.seibetsu,
        SenkyoPerson.seito,
        SenkyoPerson.genshinbetu,
        SenkyoPerson.CD_No,
        SenkyoPerson.fixedcutID,
        SenkyoPerson.MenName,
        SenkyoPerson.store_date,
        SenkyoPerson.operater,
        SenkyoPerson.output_Flg,
        SenkyoPerson.created_at,
        SenkyoPerson.updated_at,
    )

    if res["personID"]:
        person_id = _to_int_or_none(res["personID"])
        if person_id is not None:
            query = query.filter(SenkyoPerson.id == person_id)

    if res["syubetsu"]:
        query = query.filter(SenkyoPerson.syubetu.contains(res["syubetsu"]))
    if res["senkyoku"]:
        query = query.filter(SenkyoPerson.senkyoku.contains(res["senkyoku"]))
    if res["senkyokuNo"]:
        query = query.filter(SenkyoPerson.senkyokuNo.contains(res["senkyokuNo"]))
    if res["sendgroup"]:
        query = query.filter(SenkyoPerson.sendGroup.contains(res["sendgroup"]))
    if res["hirei"]:
        query = query.filter(SenkyoPerson.hirei.contains(res["hirei"]))
    if res["shimei"]:
        query = query.filter(SenkyoPerson.name.contains(res["shimei"]))
    if res["CD_No"]:
        query = query.filter(SenkyoPerson.CD_No.contains(res["CD_No"]))

    if res["startdate"]:
        start_date = datetime.strptime(res["startdate"], '%Y-%m-%d')
        query = query.filter(SenkyoPerson.created_at >= start_date)
    if res["enddate"]:
        end_date = datetime.strptime(res["enddate"], '%Y-%m-%d')
        query = query.filter(SenkyoPerson.created_at <= end_date)

    pagination = query.order_by(SenkyoPerson.id.asc()).paginate(page=page, per_page=per_page, error_out=False)
    person_rows = pagination.items

    person_columns = [
        "人物ID",
        "種別",
        "選挙区",
        "選挙区番号",
        "配信グループ",
        "比例選挙区",
        "候補者氏名",
        "ふりがな",
        "性別",
        "政党",
        "現新別",
        "CD_No",
        "固定カットID",
        "登録先面名",
        "保存日時",
        "操作担当者",
        "選挙出力フラグ",
        "作成日時",
        "更新日時",
    ]

    return render_template(
        'senkyo_person_table.html',
        person_rows=person_rows,
        person_columns=person_columns,
        savelist=savelist,
        pagination=pagination,
    )


@app.route('/senkyo_person_table_detail/<int:id>', methods=['GET', 'POST'])
def senkyo_person_table_detail(id):
    person = db.session.query(SenkyoPerson).filter(SenkyoPerson.id == id).first()
    if person is None:
        return render_template('error_404.html'), 404

    if request.method == 'POST':
        invalid = 0
        changed = False
        allow_fixedcut_insert = request.form.get('allow_fixedcut_insert') == '1'

        fixedcut_id_input = _normalize_fixedcut_id(request.form.get('fixedcutID', ''))
        if fixedcut_id_input != '' and person.fixedcutID != fixedcut_id_input:
            person.fixedcutID = fixedcut_id_input
            changed = True

        men_name_input = _to_text(request.form.get('MenName', ''))
        if men_name_input != '' and person.MenName != men_name_input:
            person.MenName = men_name_input
            changed = True

        operater_input = _to_text(request.form.get('operater', ''))
        if operater_input != '' and person.operater != operater_input:
            person.operater = operater_input
            changed = True

        output_flg_input = request.form.get('output_Flg') == '1'
        current_output_flg = bool(person.output_Flg)
        if current_output_flg != output_flg_input:
            person.output_Flg = output_flg_input
            changed = True

        store_date_text = _to_text(request.form.get('store_date', ''))
        if store_date_text != '':
            parsed_store_date = _parse_datetime_local_or_none(store_date_text)
            if parsed_store_date is None:
                invalid += 1
            elif person.store_date != parsed_store_date:
                person.store_date = parsed_store_date
                changed = True

        fixedcut_id_for_sync = _normalize_fixedcut_id(request.form.get('fixedcutID', ''))
        if fixedcut_id_for_sync == '':
            fixedcut_id_for_sync = _to_text(person.fixedcutID)

        existing_fixedcut_count = 0
        inserted_fixedcut = 0
        if fixedcut_id_for_sync:
            exists = db.session.query(FixedCut.id).filter(FixedCut.id == fixedcut_id_for_sync).first() is not None
            if exists:
                existing_fixedcut_count = 1
            elif allow_fixedcut_insert:
                db.session.add(FixedCut(
                    id=fixedcut_id_for_sync,
                    midashi=_to_text(person.name),
                    Str='',
                    colorUrl='',
                    monoUrl='',
                    GWFlg=False,
                    prodFlg=bool(_to_text(person.MenName) and _to_text(person.operater) and person.store_date is not None),
                    OTFlg=False,
                    comment='選挙顔写真固定カット',
                ))
                inserted_fixedcut = 1

        try:
            db.session.commit()
            flash(f'SenkyoPerson 差分更新完了: {1 if changed else 0}件')
            if invalid > 0:
                flash(f'日時形式が不正で更新しなかった件数: {invalid}件')
            if existing_fixedcut_count:
                flash('FixedCut既存ID: 1件')
            if inserted_fixedcut:
                flash('FixedCut新規追加: 1件')
        except Exception as exc:
            db.session.rollback()
            flash(f'※SenkyoPerson差分更新に失敗しました※ {exc}')

        person = db.session.query(SenkyoPerson).filter(SenkyoPerson.id == id).first()

    return render_template('senkyo_person_table_detail.html', person=person)


@app.route('/senkyo_person_delete/<int:id>', methods=['POST'])
def senkyo_person_delete(id):
    person = db.session.query(SenkyoPerson).filter(SenkyoPerson.id == id).first()
    if person is None:
        flash('※削除対象のSenkyoPersonレコードが見つかりません※')
        return redirect(url_for('senkyo_person_table'))

    delete_fixedcut_with_person = request.form.get('delete_fixedcut_with_person') == '1'
    fixedcut_id = _to_text(person.fixedcutID)

    try:
        deleted_fixedcut = False
        if delete_fixedcut_with_person and fixedcut_id:
            fixedcut = db.session.query(FixedCut).filter(FixedCut.id == fixedcut_id).first()
            if fixedcut is not None:
                db.session.delete(fixedcut)
                deleted_fixedcut = True

        db.session.delete(person)
        db.session.commit()
        flash(f'SenkyoPerson {id} を削除しました')
        if deleted_fixedcut:
            flash(f'FixedCut {fixedcut_id} も削除しました')
    except Exception as exc:
        db.session.rollback()
        flash(f'※SenkyoPerson削除に失敗しました※ {exc}')

    return redirect(url_for('senkyo_person_table'))


@app.route('/senkyo_sendgroup_table')
def senkyo_sendgroup_table():
    sendgroup_rows = db.session.query(SenkyoSendGroup).order_by(SenkyoSendGroup.id.asc()).all()
    sendgroup_columns = [c.name for c in SenkyoSendGroup.__table__.columns]

    return render_template(
        'senkyo_sendgroup_table.html',
        sendgroup_rows=sendgroup_rows,
        sendgroup_columns=sendgroup_columns,
    )


@app.route('/senkyo_serch', methods=['POST'])
def senkyo_serch():
    print("データを受け取りました")
    req = request.form["ID"]
    print(req)
    return f'POSTdata:{req}' 


@app.route('/general', methods=['GET', 'POST'])
@app.route('/general/<int:page>', methods=['GET', 'POST'])
def general(page=1):
    from fixedcut_app.views_general import general_handler
    return general_handler(page=page)


@app.route('/general_add', methods=['GET', 'POST'])
def general_add():
    if request.method == 'GET':
        savelist = ["","","","","","","","",""]
        return render_template('general_add.html', savelist=savelist)
    
    # フォームデータを取得
    form_id = _normalize_fixedcut_id(request.form.get("id"))
    form_midashi = request.form.get("midashi")
    form_Str = request.form.get("Str")
    form_colorUrl = request.files.get("colorUrl")
    form_monoUrl = request.files.get("monoUrl")
    form_GWFlg = request.form.get("GWFlg") == "on"
    form_prodFlg = request.form.get("prodFlg") == "on"
    form_OTFlg = request.form.get("OTFlg") == "on"
    form_comment = request.form.get("comment")

    savelist = [form_id, form_midashi, form_Str, form_colorUrl, form_monoUrl,
                form_GWFlg, form_prodFlg, form_OTFlg, form_comment]

    # 固定カットIDは必須
    if form_id == "":
        flash("※固定カットIDは必須です※")
        flash("固定カットIDを入力してから追加してください")
        return render_template('general_add.html', savelist=savelist)

    # 許可する拡張子
    ALLOWED_EXTENSIONS = {".eps", ".jpg", ".jpeg", ".svg"}

    # 拡張子チェック関数
    def is_valid_file(file):
        if not file or file.filename == "":
            return True  # ファイルなしはOK
        ext = pathlib.Path(file.filename).suffix.lower()
        return ext in ALLOWED_EXTENSIONS

    # 拡張子検証
    color_valid = is_valid_file(form_colorUrl)
    mono_valid = is_valid_file(form_monoUrl)

    # 無効な拡張子の場合、エラーメッセージを表示して終了
    if not color_valid or not mono_valid:
        flash("※画像ファイルが登録対象外の拡張子です※")
        flash('".eps", ".jpg", ".jpeg", ".svg"のみが登録できます')
        
        if form_colorUrl and form_colorUrl.filename and not color_valid:
            flash(f"カラー画像ファイル：{form_colorUrl.filename}")
        if form_monoUrl and form_monoUrl.filename and not mono_valid:
            flash(f"モノクロ画像ファイル：{form_monoUrl.filename}")
        
        app.logger.warning("Invalid file extension attempt by %s", request.remote_addr)
        return render_template('general_add.html', savelist=savelist)

    # ここから先は有効な拡張子のみ
    try:
        # ディレクトリ作成
        os.makedirs(f'fixedcut_app/templates/static/img/{form_id}/color', exist_ok=True)
        os.makedirs(f'fixedcut_app/templates/static/img/{form_id}/mono', exist_ok=True)

        # ファイル保存
        color_name = form_colorUrl.filename if form_colorUrl and form_colorUrl.filename else ""
        mono_name = form_monoUrl.filename if form_monoUrl and form_monoUrl.filename else ""

        if color_name:
            color_path = pathlib.Path(f'fixedcut_app/templates/static/img/{form_id}/color', color_name)
            form_colorUrl.save(color_path)

        if mono_name:
            mono_path = pathlib.Path(f'fixedcut_app/templates/static/img/{form_id}/mono', mono_name)
            form_monoUrl.save(mono_path)

        # DB登録
        fixedcut = FixedCut(
            id=form_id,
            midashi=form_midashi,
            Str=form_Str,
            colorUrl=f'img/{form_id}/color/{color_name}' if color_name else "",
            monoUrl=f'img/{form_id}/mono/{mono_name}' if mono_name else "",
            GWFlg=form_GWFlg,
            prodFlg=form_prodFlg,
            OTFlg=form_OTFlg,
            comment=form_comment,
        )
        db.session.add(fixedcut)

        # m_jyochu_image_cnv は fixed_cut_img_explanation を基準に分岐
        matched_jyochu = db.session.query(MJyochuImageCnv).filter(
            MJyochuImageCnv.fixed_cut_img_explanation == (form_Str or "")
        ).first()

        if matched_jyochu is None:
            new_jyochu = MJyochuImageCnv(
                fixed_cut_id=form_id,
                fixed_cut_img_explanation=form_Str or "",
            )
            db.session.add(new_jyochu)
        else:
            matched_jyochu.fixed_cut_img_explanation = form_Str or ""

        db.session.commit()

        flash(f"{form_id}をレコード追加しました！")
        app.logger.info("Added record id=%s by %s", form_id, request.remote_addr)

    except Exception as e:
        # クリーンアップ
        try:
            shutil.rmtree(f'fixedcut_app/templates/static/img/{form_id}/')
        except:
            pass

        print(f"例外args: {e.args}")
        
        if form_id == "":
            flash("※固定カットIDが入力されていません※")
            flash("固定カットIDを入力してレコード追加してください")
        else:
            flash("※すでに登録済みの固定カットIDです※")
            flash("固定カットIDを変更してレコード追加してください")
        
        app.logger.warning("Failed to add record id=%s by %s", form_id, request.remote_addr)

    return render_template('general_add.html', savelist=savelist)


@app.route('/general_detail/<string:id>', methods=['GET', 'POST'])
def general_detail(id):
    if request.method == 'GET':
        results = db.session.query(FixedCut).filter(FixedCut.id == id).all()
        if not results:
            return render_template('error_404.html'), 404
        print(results[0].GWFlg)
        jyochu_result = db.session.query(MJyochuImageCnv).filter(MJyochuImageCnv.fixed_cut_id == id).first()
        return render_template('general_detail.html', results=results, jyochu_result=jyochu_result)
    
    if request.method == 'POST':
        # フォームデータを取得（IDは変更しない）
        form_midashi = request.form.get("midashi")
        form_Str = request.form.get("Str")
        form_GWFlg = request.form.get("GWFlg") == "on"
        form_prodFlg = request.form.get("prodFlg") == "on"
        form_OTFlg = request.form.get("OTFlg") == "on"
        form_comment = request.form.get("comment")

        try:
            # 既存レコードを取得
            fixedcut = db.session.query(FixedCut).filter(FixedCut.id == id).first()
            if not fixedcut:
                flash("※指定されたレコードが見つかりません※")
                results = db.session.query(FixedCut).filter(FixedCut.id == id).all()
                jyochu_result = db.session.query(MJyochuImageCnv).filter(MJyochuImageCnv.fixed_cut_id == id).first()
                return render_template('general_detail.html', results=results, jyochu_result=jyochu_result)

            # レコード更新（ID以外）
            fixedcut.midashi = form_midashi
            fixedcut.Str = form_Str
            fixedcut.GWFlg = form_GWFlg
            fixedcut.prodFlg = form_prodFlg
            fixedcut.OTFlg = form_OTFlg
            fixedcut.comment = form_comment

            # 同一固定カットIDが m_jyochu_image_cnv に存在する場合は、説明を一体化時文字列で同期
            jyochu = db.session.query(MJyochuImageCnv).filter(MJyochuImageCnv.fixed_cut_id == id).first()
            if jyochu:
                jyochu.fixed_cut_img_explanation = form_Str or ""

            db.session.commit()

            flash(f"{id}をレコード更新しました！")
            app.logger.info("Updated record id=%s by %s", id, request.remote_addr)

        except Exception as e:
            print(f"例外args: {e.args}")
            
            flash("※レコード更新に失敗しました※")
            flash("入力内容を確認してください")
            
            app.logger.warning("Failed to update record id=%s by %s", id, request.remote_addr)

        # 更新後のデータを再取得して表示
        results = db.session.query(FixedCut).filter(FixedCut.id == id).all()
        jyochu_result = db.session.query(MJyochuImageCnv).filter(MJyochuImageCnv.fixed_cut_id == id).first()
        return render_template('general_detail.html', results=results, jyochu_result=jyochu_result)



@app.route('/general_delete/<string:id>')
def general_delete(id):
    print(f"DEBUG: general_delete called with id={id}")
    try:
        # 既存レコードを取得
        fixedcut = db.session.query(FixedCut).filter(FixedCut.id == id).first()
        print(f"DEBUG: Found record: {fixedcut}")
        if not fixedcut:
            flash("※指定されたレコードが見つかりません※")
            return redirect(url_for('general'))

        # 画像ファイルとフォルダの削除
        try:
            # IDと同名のフォルダを丸ごと削除（colorとmonoフォルダを含む）
            base_dir = pathlib.Path(f'fixedcut_app/templates/static/img/{id}')
            print(f"DEBUG: Attempting to delete directory: {base_dir}")
            print(f"DEBUG: Directory exists: {base_dir.exists()}")
            
            if base_dir.exists():
                print(f"DEBUG: Directory contents before deletion:")
                try:
                    for item in base_dir.iterdir():
                        print(f"DEBUG:   {item}")
                except Exception as e:
                    print(f"DEBUG: Error listing contents: {e}")
                
                print(f"DEBUG: Starting rmtree...")
                shutil.rmtree(base_dir)
                print(f"DEBUG: rmtree completed")
                
                # 削除確認
                if not base_dir.exists():
                    print(f"DEBUG: SUCCESS - Directory deleted: {base_dir}")
                    app.logger.info(f"Deleted image directory: {base_dir}")
                else:
                    print(f"DEBUG: ERROR - Directory still exists after rmtree: {base_dir}")
                    app.logger.error(f"Failed to delete directory after rmtree: {base_dir}")
            else:
                print(f"DEBUG: Directory does not exist: {base_dir}")
        except Exception as e:
            print(f"DEBUG: Exception during directory deletion: {e}")
            print(f"DEBUG: Exception type: {type(e)}")
            import traceback
            print(f"DEBUG: Traceback: {traceback.format_exc()}")
            app.logger.warning(f"Failed to delete image directory {id}: {e}")

        # DBからレコード削除
        print("DEBUG: Deleting record from DB")
        db.session.delete(fixedcut)
        db.session.commit()
        print("DEBUG: Committed deletion")

        # 削除後の確認
        deleted_check = db.session.query(FixedCut).filter(FixedCut.id == id).first()
        if deleted_check:
            flash("※レコード削除に失敗しました※")
            app.logger.error("Record still exists after delete attempt: %s", id)
            print("DEBUG: Record still exists after deletion!")
        else:
            flash(f"レコード {id} を削除しました")
            app.logger.info("Successfully deleted record id=%s by %s", id, request.remote_addr)
            print(f"DEBUG: Successfully deleted record {id}")

    except Exception as e:
        print(f"DEBUG: Exception in delete: {e}")
        print(f"例外args: {e.args}")
        flash("※レコード削除に失敗しました※")
        app.logger.warning("Failed to delete record id=%s by %s", id, request.remote_addr)

    # 削除後は1ページ目にリダイレクト
    print("DEBUG: Redirecting to general page 1")
    return redirect(url_for('general', page=1))



@app.route('/download/<path:filename>')
def download(filename):
    """
    filename には 'img/eeee/color/xxx.eps' のような
    static に対する相対パスが渡ってくる想定。
    """
    # セキュリティのためにパスが static_folder の外に出ないか
    # 必要なら検証を入れる（省略可）

    return send_from_directory(
        app.static_folder,      # __init__.py で static_folder='./templates/static' に設定済
        filename,
        as_attachment=True
    )

@app.errorhandler(404)
def not_found(error):
    flash(error)
    return render_template("error_404.html"), 404


@app.route('/api/check_id/<string:id>')
def check_id(id):
    """
    IDが存在するかチェックするAPI
    """
    record = db.session.query(FixedCut).filter(FixedCut.id == id).first()
    if record:
        return {'exists': True}
    else:
        return {'exists': False}


@app.route('/api/check_fixedcut_ids', methods=['POST'])
def check_fixedcut_ids():
    payload = request.get_json(silent=True) or {}
    ids = payload.get('ids') or []

    normalized_ids = []
    seen = set()
    for value in ids:
        text = _to_text(value)
        if text and text not in seen:
            normalized_ids.append(text)
            seen.add(text)

    if not normalized_ids:
        return {'existing_ids': [], 'missing_ids': []}

    existing_rows = db.session.query(FixedCut.id).filter(FixedCut.id.in_(normalized_ids)).all()
    existing_set = {row[0] for row in existing_rows}

    existing_ids = [x for x in normalized_ids if x in existing_set]
    missing_ids = [x for x in normalized_ids if x not in existing_set]
    return {'existing_ids': existing_ids, 'missing_ids': missing_ids}


@app.before_request
def log_request():
    app.logger.info("Request: %s %s from %s", request.method, request.path, request.remote_addr)